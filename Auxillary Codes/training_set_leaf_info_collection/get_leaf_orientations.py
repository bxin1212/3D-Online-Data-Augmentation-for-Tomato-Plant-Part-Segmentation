#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Apr 14 19:49:32 2022

@author: bxin
"""

import numpy as np
import pandas as pd
import os
import open3d as o3d
import h5py
from matplotlib import pyplot as plt
from sklearn.decomposition import PCA
import openpyxl as op


if __name__ == "__main__":
    
    pc_file_path = "../../dataset/inner_point_removal/Pos_Norm_Color_Sem Ins_Test DS (45 pcs)/training_set.h5"
    leaf_info_file_path = "./training_set_leaf_info.xlsx"
    
    # pc_id = 0
    
    f = h5py.File(pc_file_path, "r")
    file_names = list(f.keys())
    # print(file_names)
    
    df = pd.read_excel(leaf_info_file_path, header=0, index_col=0)
    
    # file_name = file_names[pc_id]
    file_name = "Harvest_02_PotNr_166"
    # pc_id = list(f.keys()).index(file_name)
    save_excel_col = 41
    
    pc = f.get(file_name)[()]
    
    n_leaves = np.max(pc[:, -2]) - 5  # Num of leaves includes the first small leaf containing only three leaflets. Note that some point clouds do not have this leaf, but is still counted here.
    
    for leaf_id in range(1, int(n_leaves) + 1):
        
        pts_leaf_stemwork = pc[(pc[:, -2] == leaf_id + 6) & (pc[:, -1] == 2), :]
        
        if pts_leaf_stemwork.shape[0] == 0:
            continue
        
        """
        ### To use PCA toolkit ###
        # Get the main axis of the leaf using PCA
        pca_fits = PCA(n_components=3).fit(pts_leaf_stemwork[:, 0:3])
        pca_weights = pca_fits.components_.T  # Each column refers to the weights of individual features for the specific pca component
        main_axis = pca_weights[:, 0]
        
        # Visualisation of the rotation (PCA) axis
        plt.figure()
        ax = plt.axes(projection='3d')
        ax.plot3D(pts_leaf_stemwork[:, 0], pts_leaf_stemwork[:, 1], pts_leaf_stemwork[:, 2], '.y')
        ax.plot3D(np.array([0, main_axis[0]]) + df.loc["Leaf_" + leaf_id_str + "_base_X", file_name], 
                  np.array([0, main_axis[1]]) + df.loc["Leaf_" + leaf_id_str + "_base_Y", file_name], 
                  np.array([0, main_axis[2]]) + df.loc["Leaf_" + leaf_id_str + "_base_Z", file_name], 'r')
        plt.show()
        
        """
        
        ### To use Open3D package ###
        # Create Open3d PointCloud instance with the specified leaf instance
        pc_o3d = o3d.geometry.PointCloud()
        pc_o3d.points = o3d.utility.Vector3dVector(pts_leaf_stemwork[:, 0:3])
        
        # Get leaf orientation
        pc_o3d_oriented_bounding_box = o3d.geometry.OrientedBoundingBox.create_from_points(pc_o3d.points)
        R = pc_o3d_oriented_bounding_box.R  # Three principle components?
        
        # Visualisation of the rotation (PCA) axis
        if leaf_id < 10:
            leaf_id_str = "0" + str(leaf_id)
        else:
            leaf_id_str = str(leaf_id)
        plt.figure()
        ax = plt.axes(projection='3d')
        ax.plot3D(pts_leaf_stemwork[:, 0], pts_leaf_stemwork[:, 1], pts_leaf_stemwork[:, 2], '.y')
        ax.plot3D(np.array([0, R[0, 0]/1.7]) + df.loc["Leaf_" + leaf_id_str + "_base_X", file_name], 
                  np.array([0, R[1, 0]/1.7]) + df.loc["Leaf_" + leaf_id_str + "_base_Y", file_name], 
                  np.array([0, R[2, 0]/1.7]) + df.loc["Leaf_" + leaf_id_str + "_base_Z", file_name], 'r', linewidth=3)
        ax.axis("off")
        plt.show()
        
        direction_inverse = input("To get the inverse direction of the leaf main axis, press y; otherwise, n:")
        
        if direction_inverse == 'y':
            leaf_main_axis = np.array([[0, -R[0, 0]], [0, -R[1, 0]], [0, -R[2, 0]]])
        else:
            leaf_main_axis = np.array([[0, R[0, 0]], [0, R[1, 0]], [0, R[2, 0]]])
        
        plt.figure()
        ax = plt.axes(projection='3d')
        ax.plot3D(pts_leaf_stemwork[:, 0], pts_leaf_stemwork[:, 1], pts_leaf_stemwork[:, 2], '.y')
        ax.plot3D(np.array(leaf_main_axis[0] + df.loc["Leaf_" + leaf_id_str + "_base_X", file_name]), 
                  np.array(leaf_main_axis[1] + df.loc["Leaf_" + leaf_id_str + "_base_Y", file_name]), 
                  np.array(leaf_main_axis[2] + df.loc["Leaf_" + leaf_id_str + "_base_Z", file_name]), 'r')
        plt.show()
        
        # Write the axis orientation of the leaf into the Excel file
        wb = op.load_workbook(leaf_info_file_path)
        ws = wb.active
        ws.cell(row = (leaf_id) * 6 + 5, column=save_excel_col).value = leaf_main_axis[0, 1]
        ws.cell(row = (leaf_id) * 6 + 6, column=save_excel_col).value = leaf_main_axis[1, 1]
        ws.cell(row = (leaf_id) * 6 + 7, column=save_excel_col).value = leaf_main_axis[2, 1]
        wb.save(leaf_info_file_path)
    
    # Visualisation
    # pc_visualisation = o3d.geometry.PointCloud()
    # pc_visualisation.points = o3d.utility.Vector3dVector(pts_leaf_stemwork[:, 0:3])
    # color = np.array([0., 1., 0.]) * np.ones((pts_leaf_stemwork.shape[0], 3))
    # pc_visualisation.colors = o3d.utility.Vector3dVector(color)
    # o3d.visualization.draw_geometries([pc_visualisation])